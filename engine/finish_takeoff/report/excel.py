# -*- coding: utf-8 -*-
"""물량표 Excel 출력 (7단계)."""
from __future__ import annotations

from typing import Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..takeoff.rules import FinishKind, RoomTakeoff, TakeoffSettings, summarize

KIND_LABEL = {
    FinishKind.SHEET_FLOOR: "장판/마루",
    FinishKind.FLOOR_TILE: "바닥타일",
    FinishKind.WALLPAPER: "벽 도배",
    FinishKind.CEILING_PAPER: "천장 도배",
    FinishKind.BASEBOARD: "걸레받이",
}

_HDR = PatternFill("solid", fgColor="1F4E79")
_APPROX = PatternFill("solid", fgColor="FFF2CC")


def _header(ws, row: int, cols: Sequence[str]) -> None:
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, widths: Sequence[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write(
    path: str,
    rooms: Sequence[RoomTakeoff],
    *,
    settings: Optional[TakeoffSettings] = None,
    title: str = "마감 물량 산출서",
) -> str:
    """
    실별 상세 + 마감재별 집계 시트를 쓴다.

    근사(래스터) 추적 실은 **행 전체를 음영 처리하고 비고에 명시**한다.
    검측 시 근거를 대야 하므로 숨기면 안 된다.
    """
    st = settings or TakeoffSettings()
    wb = Workbook()

    # ── 실별 상세 ──
    ws = wb.active
    ws.title = "실별 상세"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"천장고 기본 {st.ceiling_height_mm:.0f}mm · "
                f"할증 장판 {st.waste_sheet_floor:.0%} / 타일 {st.waste_floor_tile:.0%} / "
                f"도배 {st.waste_wallpaper:.0%} / 걸레받이 {st.waste_baseboard:.0%}")
    ws["A2"].font = Font(size=9, color="808080")

    cols = ["실명", "면적(㎡)", "평", "둘레(m)", "천장고(mm)",
            "장판 원(㎡)", "장판 할증(㎡)", "타일(장)",
            "벽도배 원(㎡)", "벽도배 할증(㎡)",
            "천장도배 할증(㎡)", "걸레받이(m)", "비고"]
    _header(ws, 4, cols)

    r = 5
    for rm in rooms:
        g = rm.get
        vals = [
            rm.room_name, round(rm.area_m2, 2), round(rm.pyeong, 2),
            round(rm.perimeter_m, 2), rm.ceiling_height_mm,
            round(g(FinishKind.SHEET_FLOOR).raw, 2) if g(FinishKind.SHEET_FLOOR) else None,
            round(g(FinishKind.SHEET_FLOOR).with_waste, 2) if g(FinishKind.SHEET_FLOOR) else None,
            g(FinishKind.FLOOR_TILE).count if g(FinishKind.FLOOR_TILE) else None,
            round(g(FinishKind.WALLPAPER).raw, 2) if g(FinishKind.WALLPAPER) else None,
            round(g(FinishKind.WALLPAPER).with_waste, 2) if g(FinishKind.WALLPAPER) else None,
            round(g(FinishKind.CEILING_PAPER).with_waste, 2) if g(FinishKind.CEILING_PAPER) else None,
            round(g(FinishKind.BASEBOARD).with_waste, 2) if g(FinishKind.BASEBOARD) else None,
            "근사추적(래스터) — 검측 시 확인 필요" if rm.is_approximate else "",
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            if rm.is_approximate:
                cell.fill = _APPROX
        r += 1

    ws.cell(row=r, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=r, column=2, value=round(sum(x.area_m2 for x in rooms), 2)).font = Font(bold=True)
    _autosize(ws, [22, 10, 8, 10, 11, 12, 13, 9, 13, 14, 15, 12, 30])

    # ── 마감재별 집계 ──
    ws2 = wb.create_sheet("마감재별 집계")
    _header(ws2, 1, ["마감재", "원 수량", "할증 적용", "단위", "정수수량", "할증률"])
    agg = summarize(list(rooms))
    r = 2
    for kind, line in agg.items():
        ws2.cell(row=r, column=1, value=KIND_LABEL[kind])
        ws2.cell(row=r, column=2, value=round(line.raw, 2))
        ws2.cell(row=r, column=3, value=round(line.with_waste, 2))
        ws2.cell(row=r, column=4, value=line.unit)
        ws2.cell(row=r, column=5, value=line.count)
        ws2.cell(row=r, column=6, value=f"{line.waste_rate:.1%}")
        r += 1
    _autosize(ws2, [14, 12, 12, 8, 10, 9])

    approx = sum(1 for x in rooms if x.is_approximate)
    if approx:
        ws2.cell(row=r + 1, column=1,
                 value=f"※ 근사추적 실 {approx}개 포함 — 검측 전 확인 필요").font = Font(
            bold=True, color="C00000")

    wb.save(path)
    return path
