# -*- coding: utf-8 -*-
"""
기성 청구서 Excel 출력 (5시트).

  1. 기성 요약     — 그대로 청구서에 붙일 수 있는 형태
  2. 공종별 상세   — 타입별 단위물량 · 세대수 · 금회 · 누계
  3. 세대별 내역   — **검측 근거 원장. 가장 중요.**
  4. 동별 집계     — 발주·현장 확인용
  5. 전회 대비     — 금회 증감분만 (감독 제출용)
"""
from __future__ import annotations

from collections import defaultdict
from typing import Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..billing.progress import BillingResult, WorkType, DEFAULT_WORK_TYPES
from ..billing.validator import ValidationReport

_HDR = PatternFill("solid", fgColor="1F4E79")
_SUB = PatternFill("solid", fgColor="DDEBF7")
_WARN = PatternFill("solid", fgColor="FFF2CC")
_ERR = PatternFill("solid", fgColor="FFC7CE")


def _header(ws, row: int, cols: Sequence[str]) -> None:
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _widths(ws, ws_widths: Sequence[int]) -> None:
    for i, w in enumerate(ws_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write(
    path: str,
    result: BillingResult,
    *,
    works: Sequence[WorkType] = DEFAULT_WORK_TYPES,
    validation: Optional[ValidationReport] = None,
    project_name: str = "",
    unit_price: Optional[dict[str, float]] = None,
) -> str:
    """
    기성 청구서를 쓴다.

    Args:
        path: 저장 경로.
        result: compute_billing() 결과.
        works: 공종 목록.
        validation: 검증 결과 (있으면 요약 시트에 표시).
        project_name: 현장명.
        unit_price: 공종코드 → 단가. 주면 금액 열이 붙는다.

    Returns:
        저장 경로.
    """
    wb = Workbook()
    p = result.period
    prices = unit_price or {}
    wmap = {w.code: w for w in works}

    # ── 1. 기성 요약 ────────────────────────────────────
    ws = wb.active
    ws.title = "기성 요약"
    ws["A1"] = f"{project_name} 기성 청구서" if project_name else "기성 청구서"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = (f"{p.seq}차 · {p.title} · 마감일 {p.cutoff_date:%Y-%m-%d}"
                if p else "")
    ws["A2"].font = Font(size=10, color="606060")
    if result.prev_period:
        ws["A3"] = f"전회: {result.prev_period.seq}차 ({result.prev_period.title})"
        ws["A3"].font = Font(size=9, color="808080")

    cols = ["공종", "계약물량", "전회 누계", "금회", "금회 누계", "잔여", "기성률(%)"]
    if prices:
        cols += ["단가", "금회 금액"]
    _header(ws, 5, cols)

    agg = result.by_work()
    r = 6
    tot_amount = 0.0
    for code, d in agg.items():
        w = wmap.get(code)
        contract = d["계약"]
        cum = d["누계"]
        cur = d["금회"]
        rate = (cum / contract * 100) if contract else 0.0
        vals = [w.name if w else code, round(contract, 2), round(cum - cur, 2),
                round(cur, 2), round(cum, 2), round(d["잔여"], 2), round(rate, 1)]
        if prices:
            up = prices.get(code, 0.0)
            amt = cur * up
            tot_amount += amt
            vals += [up, round(amt)]
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v)
        r += 1

    ws.cell(row=r, column=1, value="합계").font = Font(bold=True)
    if prices:
        c = ws.cell(row=r, column=len(cols), value=round(tot_amount))
        c.font = Font(bold=True)
    _widths(ws, [14, 12, 12, 12, 12, 12, 10, 12, 14])

    if validation:
        r += 2
        ws.cell(row=r, column=1, value="검증 결과").font = Font(bold=True)
        ws.cell(row=r, column=2, value=validation.summary())
        r += 1
        for iss in (validation.errors + validation.warnings)[:30]:
            c = ws.cell(row=r, column=1, value=str(iss))
            c.fill = _ERR if iss.severity.value == "error" else _WARN
            r += 1

    # ── 2. 공종별 상세 ──────────────────────────────────
    ws2 = wb.create_sheet("공종별 상세")
    _header(ws2, 1, ["공종", "세대타입", "단위물량", "세대수", "금회", "금회 누계", "잔여"])
    grp: dict[tuple[str, str], list] = defaultdict(list)
    for l in result.lines:
        grp[(l.work.code, l.unit.unit_type)].append(l)
    r = 2
    for (code, utype), lines in sorted(grp.items()):
        w = wmap.get(code)
        ws2.cell(row=r, column=1, value=w.name if w else code)
        ws2.cell(row=r, column=2, value=utype)
        ws2.cell(row=r, column=3, value=round(lines[0].unit_qty, 2))
        ws2.cell(row=r, column=4, value=len(lines))
        ws2.cell(row=r, column=5, value=round(sum(x.current_qty for x in lines), 2))
        ws2.cell(row=r, column=6, value=round(sum(x.cum_qty for x in lines), 2))
        ws2.cell(row=r, column=7, value=round(sum(x.remain_qty for x in lines), 2))
        r += 1
    _widths(ws2, [14, 12, 12, 9, 12, 13, 12])

    # ── 3. 세대별 내역 (검측 근거 원장) ─────────────────
    ws3 = wb.create_sheet("세대별 내역")
    ws3["A1"] = "검측 근거 원장 — 감독 제출용"
    ws3["A1"].font = Font(bold=True, size=12)
    _header(ws3, 3, ["동", "층", "호", "타입", "공종",
                     "전회 진도", "금회 진도", "단위물량", "금회", "금회 누계"])
    r = 4
    for l in sorted(result.lines, key=lambda x: (x.unit.building, x.unit.floor, x.unit.unit_no, x.work.code)):
        ws3.cell(row=r, column=1, value=l.unit.building)
        ws3.cell(row=r, column=2, value=l.unit.floor)
        ws3.cell(row=r, column=3, value=l.unit.unit_no)
        ws3.cell(row=r, column=4, value=l.unit.unit_type)
        ws3.cell(row=r, column=5, value=l.work.name)
        ws3.cell(row=r, column=6, value=f"{l.prev_ratio:.0%}")
        ws3.cell(row=r, column=7, value=f"{l.cum_ratio:.0%}")
        ws3.cell(row=r, column=8, value=round(l.unit_qty, 2))
        ws3.cell(row=r, column=9, value=round(l.current_qty, 2))
        ws3.cell(row=r, column=10, value=round(l.cum_qty, 2))
        r += 1
    ws3.freeze_panes = "A4"
    _widths(ws3, [8, 6, 8, 9, 12, 10, 10, 11, 11, 12])

    # ── 4. 동별 집계 ────────────────────────────────────
    ws4 = wb.create_sheet("동별 집계")
    codes = list(agg.keys())
    _header(ws4, 1, ["동"] + [wmap[c].name if c in wmap else c for c in codes])
    r = 2
    for bld, d in sorted(result.by_building().items()):
        ws4.cell(row=r, column=1, value=bld)
        for i, c in enumerate(codes, 2):
            ws4.cell(row=r, column=i, value=round(d.get(c, 0.0), 2))
        r += 1
    _widths(ws4, [10] + [13] * len(codes))

    # ── 5. 전회 대비 (금회 증감분만) ────────────────────
    ws5 = wb.create_sheet("전회 대비")
    ws5["A1"] = "금회 증감분 — 전회 대비 변동이 있는 세대만"
    ws5["A1"].font = Font(bold=True, size=12)
    _header(ws5, 3, ["동", "층", "호", "타입", "공종", "전회", "금회", "증감"])
    r = 4
    for l in result.lines:
        if abs(l.current_qty) < 1e-9:
            continue
        ws5.cell(row=r, column=1, value=l.unit.building)
        ws5.cell(row=r, column=2, value=l.unit.floor)
        ws5.cell(row=r, column=3, value=l.unit.unit_no)
        ws5.cell(row=r, column=4, value=l.unit.unit_type)
        ws5.cell(row=r, column=5, value=l.work.name)
        ws5.cell(row=r, column=6, value=f"{l.prev_ratio:.0%}")
        ws5.cell(row=r, column=7, value=f"{l.cum_ratio:.0%}")
        ws5.cell(row=r, column=8, value=round(l.current_qty, 2))
        r += 1
    _widths(ws5, [8, 6, 8, 9, 12, 8, 8, 11])

    wb.save(path)
    return path
